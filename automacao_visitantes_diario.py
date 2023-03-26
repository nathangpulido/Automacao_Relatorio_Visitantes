import openpyxl as opxl
from openpyxl.styles import Font, Fill, PatternFill

# abre planilha excel
workbook = opxl.load_workbook('visitantes_diario.xlsx')

# seleciona a "folha" atual
sheet = workbook.active

# excluí as colunas A até S
sheet.delete_cols(1, 19)

# excluí as colunas B até N
sheet.delete_cols(2, 13)

# organiza as celulas da coluna A em ordem crescente em uma lista
lista = []
for celula in sheet['A']:
    lista.append(celula.value)
lista_ordenada = sorted(lista)

# insere os valores da lista ordenada nas células da coluna A
for item in range(len(lista_ordenada)):
    sheet.cell(row=item+1, column=1).value = lista_ordenada[item]

# percorre a lista ordenada e conta a quantidade de cada item
quantidades = {}
for empresa in lista_ordenada:
    if empresa in quantidades:
        quantidades[empresa] += 1
    else:
        quantidades[empresa] = 1
quantidades_itens = {k:v for k,v in quantidades.items()}

# cria as colunas empresa e quantidade
sheet.cell(row=1, column=4, value='EMPRESA').font = Font(bold=True)
sheet.cell(row=1, column=5, value='QUANTIDADE').font = Font(bold=True)

# escreve na planilha a empresa e sua respectivas quantidades
linha = 2
for empresa, quantidade in quantidades_itens.items():
    sheet.cell(row=linha, column=4, value=empresa)
    sheet.cell(row=linha, column=5, value=quantidade)
    linha += 1

# salva alterações na planilha
workbook.save('visitantes_diario.xlsx')









